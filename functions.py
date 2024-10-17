import pandas as pd
from openpyxl import load_workbook

### função para carregar nomes de uma planilha específica ###
def carregarNomes(nomeAba, caminhoArquivo):
    df = pd.read_excel(caminhoArquivo, sheet_name=nomeAba)
    return df.iloc[:, 0].tolist()  ### retorna a primeira coluna como uma lista de nomes ###

### função para carregar critérios e seus pontos de uma planilha específica ###
def carregarCriterios(nomeAba, caminhoArquivo):
    df = pd.read_excel(caminhoArquivo, sheet_name=nomeAba)
    return dict(zip(df['Requisito'], df['Pontos Totais']))  ### cria um dicionário com critérios e pontos ###

### função para adicionar dados à planilha na aba específica ###
def adicionarDados(caminhoArquivo, nomeAba, dados):
    wb = load_workbook(caminhoArquivo)
    planilha = wb[nomeAba]

    proximaLinha = planilha.max_row + 1  ### determina a próxima linha vazia ###
    for numeroColuna, valor in enumerate(dados, start=1):
        planilha.cell(row=proximaLinha, column=numeroColuna, value=valor)  ### insere os dados na planilha ###

    wb.save(caminhoArquivo)  ### salva as alterações no arquivo Excel ###
